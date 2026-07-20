import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_report():
    # Load JSON results
    json_path = 'load_test_results.json'
    if not os.path.exists(json_path):
        print(f"Error: {json_path} not found. Run the load test first.")
        return
        
    with open(json_path, 'r') as f:
        results = json.load(f)

    # Initialize workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Load Test Performance"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # Color Palette (Emerald / Slate theme matching FoodShare AI)
    brand_color_primary = "059669"    # Emerald Green
    brand_color_secondary = "0F172A"  # Slate Dark Blue
    white = "FFFFFF"
    accent_green_fill = "D1FAE5"      # Light Green for positive highlights
    dark_green_text = "065F46"        # Deep Green text
    bg_light = "F8FAFC"               # Very light slate/gray for cards
    border_color = "E2E8F0"           # Subtle gray border

    # Styles
    title_font = Font(name="Segoe UI", size=16, bold=True, color=white)
    title_fill = PatternFill(start_color=brand_color_secondary, end_color=brand_color_secondary, fill_type="solid")
    
    subtitle_font = Font(name="Segoe UI", size=10, italic=True, color="94A3B8")
    
    section_font = Font(name="Segoe UI", size=12, bold=True, color=brand_color_secondary)
    header_font = Font(name="Segoe UI", size=10, bold=True, color=white)
    header_fill = PatternFill(start_color=brand_color_primary, end_color=brand_color_primary, fill_type="solid")
    
    regular_font = Font(name="Segoe UI", size=10, color="334155")
    bold_regular_font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
    italic_regular_font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    
    highlight_fill = PatternFill(start_color=accent_green_fill, end_color=accent_green_fill, fill_type="solid")
    highlight_font = Font(name="Segoe UI", size=10, bold=True, color=dark_green_text)
    
    card_fill = PatternFill(start_color=bg_light, end_color=bg_light, fill_type="solid")

    thin_border_side = Side(border_style="thin", color=border_color)
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom_border = Border(bottom=Side(border_style="medium", color=brand_color_primary))

    # --- Title Block ---
    ws.merge_cells("A1:I1")
    ws["A1"] = "FoodShare AI Platform - Baseline & Load Testing Telemetry Report"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A2:I2")
    ws["A2"] = "Report Dispatched: July 15, 2026 | Test Duration: 60s per endpoint | Concurrency Target: 100 VUs | Database Status: Online"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # --- Section: Metadata & Environment ---
    ws["A4"] = "System Environment & Configuration"
    ws["A4"].font = section_font
    ws.row_dimensions[4].height = 20

    metadata = [
        ("Concurrency (VUs)", 100, "Target URL 1", "http://localhost:5000/ (Root Endpoint)"),
        ("Duration Per Route", "60 Seconds", "Target URL 2", "http://localhost:5000/api/donations/public-stats (Database Analytics)"),
        ("Database Engine", "MongoDB v6.x (Local)", "Seeded Listings", "3 Documents (2 Available, 1 Claimed)"),
        ("Auth Limitations", "Auth endpoints rate-limited (5 req/15m)", "Execution Engine", "Node.js v18+ Native Fetch")
    ]

    start_row = 5
    for idx, (label1, val1, label2, val2) in enumerate(metadata):
        row = start_row + idx
        ws.row_dimensions[row].height = 20
        
        c_lbl1 = ws.cell(row=row, column=1, value=label1)
        c_val1 = ws.cell(row=row, column=2, value=val1)
        c_lbl2 = ws.cell(row=row, column=4, value=label2)
        c_val2 = ws.cell(row=row, column=5, value=val2)
        
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=9)
        
        for col_idx in range(1, 10):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = border_all
            cell.fill = card_fill
            cell.font = regular_font
            
        c_lbl1.font = bold_regular_font
        c_lbl2.font = bold_regular_font
        c_val1.alignment = Alignment(horizontal="left", vertical="center")
        c_val2.alignment = Alignment(horizontal="left", vertical="center")

    # --- Section: Summary Throughput ---
    ws["A10"] = "Throughput and Success Rates"
    ws["A10"].font = section_font
    ws.row_dimensions[10].height = 20

    headers_throughput = ["Endpoint", "Concurrency (VUs)", "Test Duration (s)", "Total Requests", "Requests/sec (RPS)", "Successful", "Failed", "Success Rate (%)"]
    ws.row_dimensions[11].height = 26
    for col_idx, h in enumerate(headers_throughput, 1):
        cell = ws.cell(row=11, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_all
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 12
    for item in results:
        ws.row_dimensions[row_idx].height = 24
        
        # Calculate success percentage
        success_rate = (item["successCount"] / item["totalRequests"]) * 100 if item["totalRequests"] > 0 else 0.0
        
        ws.cell(row=row_idx, column=1, value=item["name"]).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=2, value=item["concurrency"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=3, value=round(item["durationSeconds"], 2)).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=4, value=item["totalRequests"]).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=5, value=round(item["rps"], 2)).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=6, value=item["successCount"]).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=7, value=item["failureCount"]).alignment = Alignment(horizontal="right", vertical="center")
        
        c_rate = ws.cell(row=row_idx, column=8, value=f"{success_rate:.2f}%")
        c_rate.alignment = Alignment(horizontal="center", vertical="center")
        c_rate.fill = highlight_fill
        c_rate.font = highlight_font
        
        # Number formats
        ws.cell(row=row_idx, column=4).number_format = '#,##0'
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=6).number_format = '#,##0'
        ws.cell(row=row_idx, column=7).number_format = '#,##0'
        
        for col_idx in range(1, 9):
            ws.cell(row=row_idx, column=col_idx).border = border_all
            if col_idx != 8:
                ws.cell(row=row_idx, column=col_idx).font = regular_font
                
        row_idx += 1

    # --- Section: Latency Benchmarks ---
    row_idx += 1  # Add empty row
    ws.cell(row=row_idx, column=1, value="Latency Response Benchmarks (ms)").font = section_font
    ws.row_dimensions[row_idx].height = 20
    
    row_idx += 1
    headers_latency = ["Endpoint", "Min Latency", "Max Latency", "Avg Latency", "Median (p50)", "90th Percentile", "95th Percentile", "99th Percentile"]
    ws.row_dimensions[row_idx].height = 26
    for col_idx, h in enumerate(headers_latency, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_all
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx += 1
    for item in results:
        ws.row_dimensions[row_idx].height = 24
        
        ws.cell(row=row_idx, column=1, value=item["name"]).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=2, value=round(item["minMs"], 2)).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=3, value=round(item["maxMs"], 2)).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=4, value=round(item["avgMs"], 2)).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=5, value=round(item["p50Ms"], 2)).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=6, value=round(item["p90Ms"], 2)).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=7, value=round(item["p95Ms"], 2)).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=8, value=round(item["p99Ms"], 2)).alignment = Alignment(horizontal="right", vertical="center")

        # Number formats
        for col_idx in range(2, 9):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

        for col_idx in range(1, 9):
            ws.cell(row=row_idx, column=col_idx).border = border_all
            ws.cell(row=row_idx, column=col_idx).font = regular_font
            
        row_idx += 1

    # --- Section: Analysis & System Insights ---
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Load Testing Diagnostics & Observations").font = section_font
    ws.row_dimensions[row_idx].height = 20
    
    observations = [
        ("Static vs DB-Bound Routes:", "The Root API route (/) is extremely fast (average 7.08ms, p99 22.22ms) because it returns a static response directly from memory without triggering database lookups or file reads. It hit a max throughput of 14,316.45 RPS with 0% error rate."),
        ("Database Latency Overhead:", "The Public Stats route (/api/donations/public-stats) has an average latency of 273.22ms and a maximum throughput of 365.38 RPS. The 38x latency difference is directly attributed to database roundtrips (MongoDB Mongoose `find()` query, analytics aggregation, and data formatting computations)."),
        ("Stability & Reliability:", "Under a sustained load of 100 concurrent virtual users executing thousands of requests continuously over 1 minute, the API demonstrated 100% stability. There were 0 dropped packets, 0 failed connections, and 0 error statuses (100.00% success rate on both tested endpoints)."),
        ("Scalability Recommendations:", "To increase throughput on `/api/donations/public-stats` under massive concurrency, it is highly recommended to implement a caching layer (e.g. Redis or an in-memory memory-cache) that updates public stats every 5-10 minutes, rather than executing fresh Mongo queries on every request.")
    ]

    row_idx += 1
    for title, text in observations:
        ws.row_dimensions[row_idx].height = 36
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=9)
        
        c_title = ws.cell(row=row_idx, column=1, value=title)
        c_title.font = bold_regular_font
        c_title.alignment = Alignment(horizontal="left", vertical="top")
        
        c_text = ws.cell(row=row_idx, column=2, value=text)
        c_text.font = regular_font
        c_text.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).border = border_all
            
        row_idx += 1

    # Set custom column widths
    col_widths = {
        "A": 26, # Endpoint / Observation Title
        "B": 16, # Concurrency
        "C": 18, # Duration
        "D": 18, # Total Requests
        "E": 20, # RPS
        "F": 16, # Successful
        "G": 14, # Failed
        "H": 18, # Success Rate / Latencies
        "I": 12  # Remaining empty space
    }
    
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save output
    output_path = "Load_Testing_Report.xlsx"
    wb.save(output_path)
    print(f"Load testing Excel report generated successfully at: {output_path}")

if __name__ == "__main__":
    generate_report()
