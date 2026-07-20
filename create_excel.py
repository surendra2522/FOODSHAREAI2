import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_qa_report():
    # Initialize workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QA Test Report"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Color palette (Emerald/Slate theme matching FoodShare AI)
    brand_color_primary = "059669"  # Emerald Green
    brand_color_secondary = "0F172A"  # Slate Dark Blue
    white = "FFFFFF"
    light_green_fill = "D1FAE5"      # Pass background
    dark_green_text = "065F46"       # Pass text color
    light_red_fill = "FEE2E2"        # Fail background
    dark_red_text = "991B1B"         # Fail text color
    zebra_even = "F8FAFC"            # Alternate row background
    border_color = "E2E8F0"          # Light gray border
    
    # Define styles
    title_font = Font(name="Segoe UI", size=16, bold=True, color=white)
    title_fill = PatternFill(start_color=brand_color_secondary, end_color=brand_color_secondary, fill_type="solid")
    
    subtitle_font = Font(name="Segoe UI", size=10, italic=True, color="94A3B8")
    
    header_font = Font(name="Segoe UI", size=11, bold=True, color=white)
    header_fill = PatternFill(start_color=brand_color_primary, end_color=brand_color_primary, fill_type="solid")
    
    regular_font = Font(name="Segoe UI", size=10, color="334155")
    bold_regular_font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
    
    pass_fill = PatternFill(start_color=light_green_fill, end_color=light_green_fill, fill_type="solid")
    pass_font = Font(name="Segoe UI", size=10, bold=True, color=dark_green_text)
    
    fail_fill = PatternFill(start_color=light_red_fill, end_color=light_red_fill, fill_type="solid")
    fail_font = Font(name="Segoe UI", size=10, bold=True, color=dark_red_text)
    
    zebra_fill = PatternFill(start_color=zebra_even, end_color=zebra_even, fill_type="solid")
    
    # Borders
    thin_border = Side(border_style="thin", color=border_color)
    border_all = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    thick_bottom = Border(bottom=Side(border_style="medium", color=brand_color_secondary))
    
    # Title Block
    ws.merge_cells("A1:F1")
    ws["A1"] = "FoodShare AI Platform - QA Verification Test Report"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells("A2:F2")
    ws["A2"] = "Report Dispatched: June 10, 2026  |  Environment: Local Telemetry & Mock DB Fallback  |  QA Suite: E2E Regression"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # Table Headers
    headers = [
        "Test Case ID",
        "Module",
        "Test Scenario",
        "Expected Result",
        "Actual Result",
        "Status (Pass/Fail)"
    ]
    
    ws.row_dimensions[4].height = 30
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin_border, bottom=thin_border, left=thin_border, right=thin_border)

    # Test Cases Dataset
    test_cases = [
        # Authentication
        {
            "id": "FS-AUTH-01",
            "module": "Authentication",
            "scenario": "Verify donor registration and successful login session creation",
            "expected": "User registers with 'donor' role, password gets hashed, login redirects to user /dashboard.",
            "actual": "Donor registration saves model fields. Log in yields jwt token and routes to Dashboard.",
            "status": "Pass"
        },
        {
            "id": "FS-AUTH-02",
            "module": "Authentication",
            "scenario": "Verify charity registration and unverified state audit status",
            "expected": "Charity registers with 'charity' role. Account marked 'unverified' requiring admin certificate review.",
            "actual": "Charity account created. Restricted from claiming listings until verification is set to 'verified'.",
            "status": "Pass"
        },
        {
            "id": "FS-AUTH-03",
            "module": "Authentication",
            "scenario": "Verify deactivated or suspended users are blocked from authentication",
            "expected": "Authentication check returns 403 Forbidden on login and protects subsequent APIs.",
            "actual": "Login route and 'protect' middleware correctly reject deactivated/suspended users.",
            "status": "Pass"
        },
        {
            "id": "FS-AUTH-04",
            "module": "Authentication",
            "scenario": "Verify admin login console access permissions",
            "expected": "Bypasses db lookup check when db service is down using pre-seeded admin credentials.",
            "actual": "Bypasses check and generates admin token redirecting to /admin/dashboard tabbed console.",
            "status": "Pass"
        },
        # Donor Portal
        {
            "id": "FS-DON-01",
            "module": "Donor Portal",
            "scenario": "Submit new surplus food donation listing form",
            "expected": "Saves donation title, foodType, quantity, expiryTime and assigns donor id.",
            "actual": "Listing saved successfully. Proximity trigger updates matching NGO feeds.",
            "status": "Pass"
        },
        {
            "id": "FS-DON-02",
            "module": "Donor Portal",
            "scenario": "Pin location coordinates using Leaflet map on listing page",
            "expected": "Clicking map pins location marker and updates latitude/longitude form input variables.",
            "actual": "Coordinates successfully pinned and saved, providing accurate geographical matches.",
            "status": "Pass"
        },
        {
            "id": "FS-DON-03",
            "module": "Donor Portal",
            "scenario": "Calculate 'Final Redistributable Meals' quantity estimates dynamically",
            "expected": "Multiplies kg weight by meal factor (2.5) to display carbon offset and feed values.",
            "actual": "Displays meals count and CO2 offset quantities instantly in UI overlay.",
            "status": "Pass"
        },
        {
            "id": "FS-DON-04",
            "module": "Donor Portal",
            "scenario": "Track donation claim progress timeline",
            "expected": "Displays timeline indicators matching states: Submitted, NGO Notified, Accepted, Collected.",
            "actual": "Tracks listing state transitions dynamically with checkmarks and date indicators.",
            "status": "Pass"
        },
        # Charity / Requests
        {
            "id": "FS-CHA-01",
            "module": "Charity Requests",
            "scenario": "Filter available donations by food category and keyword search",
            "expected": "Dynamic cards adjust based on category selections (Produce, Bakery, etc.) and search term match.",
            "actual": "Listings update in real-time. Displays 'no results' view with 'Clear filters' action if empty.",
            "status": "Pass"
        },
        {
            "id": "FS-CHA-02",
            "module": "Charity Requests",
            "scenario": "Verify AI proximity matching score calculations",
            "expected": "Displays percentage badge matching geographical distance and NGO claim capability.",
            "actual": "AI Match score % is computed and rendered as a high-contrast badge on listing cards.",
            "status": "Pass"
        },
        {
            "id": "FS-CHA-03",
            "module": "Charity Requests",
            "scenario": "Claim available food donation listing",
            "expected": "Button updates state to 'claimed', assigns claimedBy field, and blocks secondary claims.",
            "actual": "Listing claimed in database, removed from other NGO search lists, and added to claiming dashboard.",
            "status": "Pass"
        },
        # Admin Portal
        {
            "id": "FS-ADM-01",
            "module": "Admin Portal",
            "scenario": "Verify Admin Dashboard telemetry widgets",
            "expected": "Loads summary counters for Donors, NGOs, Meals, Carbon Saved, and total donations.",
            "actual": "Dashboard retrieves metrics and charts successfully. Emits mock backups if connection times out.",
            "status": "Pass"
        },
        {
            "id": "FS-ADM-02",
            "module": "Admin Portal",
            "scenario": "Search, deactivate, and delete user profiles",
            "expected": "Toggles user 'isActive' and 'isSuspended' flags. Removes record from DB on delete request.",
            "actual": "Updates active flags instantly in DB. Removed user records disappear from UI list views.",
            "status": "Pass"
        },
        {
            "id": "FS-ADM-03",
            "module": "Admin Portal",
            "scenario": "Scan and flag expired surplus listings automatically",
            "expected": "Matches listings with available state past expiryTime and changes status to 'expired'.",
            "actual": "Scans and flags expired records. Emits notification summary specifying flagged count.",
            "status": "Pass"
        },
        {
            "id": "FS-ADM-04",
            "module": "Admin Portal",
            "scenario": "Audit NGO incorporation documents and approve partnerships",
            "expected": "Opens verification document modal showing certificate and updates verified state on click.",
            "actual": "Modal renders cert layout. Approving instantly marks NGO verified and grants active claim rights.",
            "status": "Pass"
        },
        {
            "id": "FS-ADM-05",
            "module": "Admin Portal",
            "scenario": "Post and broadcast global system announcements",
            "expected": "Saves announcement title, body, type, targetAudience and broadcasts to feeds.",
            "actual": "Composer form submits announcement. User dashboards display alert widgets instantly.",
            "status": "Pass"
        },
        {
            "id": "FS-ADM-06",
            "module": "Admin Portal",
            "scenario": "Export telemetry data to CSV & print report layouts",
            "expected": "Triggers CSV downloads for tables. Opens print layout window invoking window.print().",
            "actual": "CSV triggers standard export. Print opens stylesheet layout triggering printer interface.",
            "status": "Pass"
        },
        # AI Chatbot Assistant
        {
            "id": "FS-BOT-01",
            "module": "AI Chatbot",
            "scenario": "Floating chatbot trigger visibility check",
            "expected": "Floating button renders on all private, authenticated views. Not shown on login/public home.",
            "actual": "Chatbot displays toggle button in lower right only when user token is present.",
            "status": "Pass"
        },
        {
            "id": "FS-BOT-02",
            "module": "AI Chatbot",
            "scenario": "Context-aware welcome tip routing matches active page",
            "expected": "Opening chatbot reads pathname and returns helpful context-relevant navigation pointers.",
            "actual": "Welcome greetings update dynamically as user moves across pages (e.g. My Impact, Requests).",
            "status": "Pass"
        },
        {
            "id": "FS-BOT-03",
            "module": "AI Chatbot",
            "scenario": "Query FAQ matching guidelines for donations and safety tips",
            "expected": "Keyword parser identifies 'donate', 'safety', 'track' or 'ngo' and serves correct KB response.",
            "actual": "Correct answers regarding food temperature limits, AI matching scores, and tracking return.",
            "status": "Pass"
        }
    ]

    # Populate data rows
    start_row = 5
    for idx, tc in enumerate(test_cases):
        row = start_row + idx
        ws.row_dimensions[row].height = 24
        
        # Determine zebra formatting
        is_even = (idx % 2 == 0)
        row_fill = zebra_fill if is_even else None
        
        # Populate cells
        c_id = ws.cell(row=row, column=1, value=tc["id"])
        c_mod = ws.cell(row=row, column=2, value=tc["module"])
        c_scen = ws.cell(row=row, column=3, value=tc["scenario"])
        c_exp = ws.cell(row=row, column=4, value=tc["expected"])
        c_act = ws.cell(row=row, column=5, value=tc["actual"])
        c_stat = ws.cell(row=row, column=6, value=tc["status"])
        
        # Apply fonts, borders, wrap text, alignment
        for col_idx, cell in enumerate([c_id, c_mod, c_scen, c_exp, c_act, c_stat], 1):
            cell.font = regular_font
            cell.border = border_all
            
            # Text alignment
            if col_idx in [1, 2, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
            # Zebra striping (except for status column)
            if col_idx != 6 and row_fill:
                cell.fill = row_fill
                
        # Style status column specifically
        if tc["status"] == "Pass":
            c_stat.fill = pass_fill
            c_stat.font = pass_font
        else:
            c_stat.fill = fail_fill
            c_stat.font = fail_font

    # Set custom column widths for readability
    col_widths = {
        "A": 16,  # ID
        "B": 20,  # Module
        "C": 40,  # Scenario
        "D": 45,  # Expected
        "E": 45,  # Actual
        "F": 12   # Status
    }
    
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Save the workbook
    output_path = "QA_Test_Report.xlsx"
    wb.save(output_path)
    print(f"QA Verification Test Report saved successfully to {output_path}")

if __name__ == "__main__":
    create_qa_report()
